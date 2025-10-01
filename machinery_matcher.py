def _calculate_match(self, prospect, provider, technology_filter=None):
        """Calculate match score between prospect and provider"""
        
        score = 0
        reasons = []
        
        # Technology matching (HIGH PRIORITY if filter is set)
        if technology_filter:
            prospect_processes = prospect.get('production_processes', [])
            provider_techs = provider.get('technologies', []) + provider.get('processes', [])
            tech_keywords = TECHNOLOGY_KEYWORDS.get(technology_filter, [])
            
            prospect_has_tech = any(
                any(keyword.lower() in process.lower() for keyword in tech_keywords)
                for process in prospect_processes
            ) if prospect_processes else False
            
            provider_has_tech = any(
                any(keyword.lower() in tech.lower() for keyword in tech_keywords)
                for tech in provider_techs
            ) if provider_techs else False
            
            if prospect_has_tech and provider_has_tech:
                score += 35
                reasons.append(f"Both use {technology_filter} technology")
            elif provider_has_tech:
                score += 20
                reasons.append(f"Provider specializes in {technology_filter}")
        
        """
SCALABLE MACHINERY MATCHER v2.0
Optimized for 1500+ prospects and 1900+ K2025 exhibitors

Key Optimizations:
- Batch processing with parallel execution
- Caching to avoid re-scraping
- Progressive analysis (can resume if interrupted)
- Smart sampling for faster results
- Database storage for large datasets
"""

import pandas as pd
import anthropic
import json
import re
from datetime import datetime
import requests
from bs4 import BeautifulSoup
import base64
from PIL import Image
from io import BytesIO
from urllib.parse import urljoin
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
import sqlite3
from pathlib import Path
import hashlib
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Configuration
try:
    from config import (ANTHROPIC_API_KEY, CSV_FILE_PATH, TOP_N_PROVIDERS, 
                       MAX_PROSPECTS_TO_ANALYZE, ENABLE_WEB_SCRAPING, FILTER_BY_TECHNOLOGY)
except ImportError:
    ANTHROPIC_API_KEY = None
    CSV_FILE_PATH = None
    TOP_N_PROVIDERS = 10
    MAX_PROSPECTS_TO_ANALYZE = 1500
    ENABLE_WEB_SCRAPING = True
    FILTER_BY_TECHNOLOGY = None

# K2025 Exhibitor scraping URL
K2025_SEARCH_URL = "https://www.k-online.com/vis/v1/en/search"
K2025_DIRECTORY_URL = "https://www.k-online.com/vis/v1/en/directory/{letter}"
K2025_CATALOGUE_URL = "https://www.k-online.com/vis/v1/en/catalogue"

# Technology mapping - COMPREHENSIVE with all variations and subtypes
TECHNOLOGY_KEYWORDS = {
    # INJECTION MOLDING - All variations
    'injection': [
        'injection molding', 'injection moulding', 'injection', 'IMM',
        'plastic injection', 'insert molding', 'overmolding', 'overmoulding',
        'two-shot molding', '2-shot molding', 'multi-shot', 'co-injection',
        'gas-assisted injection', 'water-assisted injection', 'foam injection',
        'thin-wall injection', 'micro injection', 'micro-molding',
        'vertical injection', 'horizontal injection', 'toggle injection',
        'hydraulic injection', 'electric injection', 'hybrid injection',
        'LSR injection', 'liquid silicone', 'rubber injection',
        'reaction injection', 'RIM', 'structural foam', 'MuCell'
    ],
    
    # EXTRUSION - All types
    'extrusion': [
        'extrusion', 'extruder', 'profile extrusion', 'pipe extrusion',
        'film extrusion', 'blown film', 'cast film', 'sheet extrusion',
        'flat die extrusion', 'coextrusion', 'co-extrusion', 'multilayer',
        'wire coating', 'cable extrusion', 'tube extrusion', 'hose extrusion',
        'compounding', 'twin-screw', 'single-screw', 'counter-rotating',
        'window profile', 'PVC extrusion', 'WPC extrusion', 'foam extrusion',
        'extrusion coating', 'extrusion lamination', 'strand pelletizing'
    ],
    
    # BLOW MOLDING - All variations
    'blow_molding': [
        'blow molding', 'blow moulding', 'blowing', 'bottle blowing',
        'extrusion blow molding', 'EBM', 'injection blow molding', 'IBM',
        'stretch blow molding', 'SBM', 'ISBM', 'injection stretch',
        'PET blow', 'preform', 'preform injection', 'bottle production',
        'container blowing', 'HDPE bottle', 'multilayer blow',
        '3D blow molding', 'shuttle blow', 'continuous blow',
        'accumulator head', 'rotary blow', 'linear blow'
    ],
    
    # THERMOFORMING
    'thermoforming': [
        'thermoforming', 'vacuum forming', 'pressure forming',
        'twin-sheet thermoforming', 'plug-assist', 'drape forming',
        'matched mold', 'forming', 'vacuum thermoform', 'blister pack',
        'clamshell', 'deep draw', 'skin packaging', 'roll-fed thermoform'
    ],
    
    # COMPRESSION MOLDING
    'compression': [
        'compression molding', 'compression moulding', 'compression press',
        'SMC', 'BMC', 'sheet molding compound', 'bulk molding compound',
        'GMT', 'glass mat thermoplastic', 'transfer molding',
        'resin transfer molding', 'RTM', 'compression thermoset',
        'rubber compression', 'silicone compression'
    ],
    
    # ROTATIONAL MOLDING
    'rotomolding': [
        'rotational molding', 'rotomolding', 'rotomoulding',
        'roto molding', 'rotocasting', 'carousel molding',
        'rock and roll', 'shuttle rotomolding', 'rotational casting'
    ],
    
    # FILM & SHEET
    'film_sheet': [
        'film production', 'film blowing', 'film casting',
        'sheet production', 'sheet casting', 'calendering',
        'film stretching', 'BOPP', 'BOPET', 'biaxial orientation',
        'monolayer film', 'multilayer film', 'barrier film',
        'stretch film', 'shrink film', 'packaging film'
    ],
    
    # FOAMING
    'foaming': [
        'foam molding', 'foam injection', 'expanded polystyrene', 'EPS',
        'XPS', 'extruded polystyrene', 'polyurethane foam', 'PU foam',
        'foam extrusion', 'microcellular foam', 'structural foam',
        'bead foaming', 'EPP', 'EPE', 'foam manufacturing'
    ],
    
    # PULTRUSION & COMPOSITES
    'composites': [
        'pultrusion', 'filament winding', 'hand layup', 'spray up',
        'resin infusion', 'vacuum infusion', 'autoclave molding',
        'prepreg', 'composite manufacturing', 'fiber reinforced',
        'glass fiber', 'carbon fiber', 'FRP', 'GFRP', 'CFRP'
    ],
    
    # 3D PRINTING / ADDITIVE
    'additive': [
        '3D printing', 'additive manufacturing', 'FDM', 'FFF',
        'SLA', 'SLS', 'stereolithography', 'selective laser',
        'material jetting', 'binder jetting', 'powder bed fusion',
        'fused deposition', 'rapid prototyping', 'AM'
    ],
    
    # WELDING & ASSEMBLY
    'welding': [
        'ultrasonic welding', 'vibration welding', 'hot plate welding',
        'laser welding', 'infrared welding', 'spin welding',
        'friction welding', 'plastic welding', 'heat staking',
        'ultrasonic insertion', 'assembly'
    ],
    
    # RECYCLING & COMPOUNDING
    'recycling': [
        'recycling', 'regranulation', 'agglomeration', 'densification',
        'wash line', 'flake production', 'pelletizing', 'reprocessing',
        'post-consumer', 'PCR', 'regrind', 'reclaim', 'circular economy',
        'mechanical recycling', 'chemical recycling', 'pyrolysis'
    ],
    
    # DECORATING & FINISHING
    'decorating': [
        'pad printing', 'screen printing', 'hot stamping', 'IML',
        'in-mold labeling', 'in-mold decoration', 'IMD', 'painting',
        'coating', 'metalizing', 'vacuum metalizing', 'chrome plating',
        'laser marking', 'engraving', 'printing on plastic'
    ]
}

# Create reverse mapping for easy lookup
TECHNOLOGY_CATEGORIES = {}
for category, keywords in TECHNOLOGY_KEYWORDS.items():
    for keyword in keywords:
        TECHNOLOGY_CATEGORIES[keyword.lower()] = category

# User-friendly technology names
TECHNOLOGY_DISPLAY_NAMES = {
    'injection': 'Injection Molding (all types)',
    'extrusion': 'Extrusion (all types)',
    'blow_molding': 'Blow Molding (PET, HDPE, all types)',
    'thermoforming': 'Thermoforming & Vacuum Forming',
    'compression': 'Compression & Transfer Molding',
    'rotomolding': 'Rotational Molding',
    'film_sheet': 'Film & Sheet Production',
    'foaming': 'Foam Manufacturing',
    'composites': 'Composites & Pultrusion',
    'additive': '3D Printing & Additive Manufacturing',
    'welding': 'Plastic Welding & Assembly',
    'recycling': 'Recycling & Compounding',
    'decorating': 'Decorating & Finishing'
}

class CacheDB:
    """SQLite cache for scraped data to avoid re-scraping"""
    
    def __init__(self, db_path="machinery_cache.db"):
        self.conn = sqlite3.connect(db_path, check_same_thread=False)
        self.setup_tables()
    
    def setup_tables(self):
        """Create cache tables"""
        self.conn.execute('''
            CREATE TABLE IF NOT EXISTS prospect_data (
                url TEXT PRIMARY KEY,
                company TEXT,
                data TEXT,
                scraped_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        self.conn.execute('''
            CREATE TABLE IF NOT EXISTS provider_data (
                name TEXT PRIMARY KEY,
                data TEXT,
                scraped_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        self.conn.execute('''
            CREATE TABLE IF NOT EXISTS k2025_exhibitors (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE,
                url TEXT,
                hall TEXT,
                stand TEXT,
                country TEXT,
                products TEXT,
                scraped_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        self.conn.commit()
    
    def get_prospect_cache(self, url):
        """Get cached prospect data"""
        cursor = self.conn.execute(
            "SELECT data FROM prospect_data WHERE url = ?", (url,)
        )
        row = cursor.fetchone()
        return json.loads(row[0]) if row else None
    
    def save_prospect_cache(self, url, company, data):
        """Save prospect data to cache"""
        self.conn.execute(
            "INSERT OR REPLACE INTO prospect_data (url, company, data) VALUES (?, ?, ?)",
            (url, company, json.dumps(data))
        )
        self.conn.commit()
    
    def get_k2025_exhibitors(self):
        """Get all cached K2025 exhibitors"""
        cursor = self.conn.execute("SELECT * FROM k2025_exhibitors")
        return cursor.fetchall()
    
    def save_k2025_exhibitor(self, name, url, hall, stand, country, products):
        """Save K2025 exhibitor"""
        self.conn.execute(
            "INSERT OR REPLACE INTO k2025_exhibitors (name, url, hall, stand, country, products) VALUES (?, ?, ?, ?, ?, ?)",
            (name, url, hall, stand, country, products)
        )
        self.conn.commit()


class K2025Scraper:
    """Scrapes K2025 exhibitor database"""
    
    def __init__(self, cache_db):
        self.cache = cache_db
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })
    
    def scrape_all_exhibitors(self, category_filter="machinery"):
        """Scrape all K2025 exhibitors - focuses on machinery/equipment"""
        
        print("\n" + "="*90)
        print("🏭 SCRAPING K2025 EXHIBITOR DATABASE")
        print("="*90)
        
        # Check cache first
        cached = self.cache.get_k2025_exhibitors()
        if cached and len(cached) > 100:
            print(f"✓ Found {len(cached)} exhibitors in cache")
            return self._format_exhibitors(cached)
        
        print("🔍 Fetching fresh data from K2025 website...")
        exhibitors = []
        
        # Method 1: Try catalogue/category approach
        exhibitors.extend(self._scrape_by_category())
        
        # Method 2: Try alphabetical directory
        if len(exhibitors) < 100:
            exhibitors.extend(self._scrape_by_directory())
        
        # Save to cache
        for exhibitor in exhibitors:
            self.cache.save_k2025_exhibitor(
                exhibitor['name'],
                exhibitor.get('url', ''),
                exhibitor.get('hall', ''),
                exhibitor.get('stand', ''),
                exhibitor.get('country', ''),
                json.dumps(exhibitor.get('products', []))
            )
        
        print(f"\n✓ Scraped {len(exhibitors)} machinery providers from K2025")
        return exhibitors
    
    def _scrape_by_category(self):
        """Scrape machinery category from K2025"""
        exhibitors = []
        
        try:
            # Try to get machinery category (category 03)
            url = "https://www.k-online.com/vis/v1/en/search?f_prod=k2025.03*"
            response = self.session.get(url, timeout=15)
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Find exhibitor cards/listings
            exhibitor_elements = soup.find_all(['div', 'article'], class_=re.compile('exhibitor|company|profile'))
            
            for elem in exhibitor_elements[:100]:  # Limit initial scrape
                try:
                    name = elem.find(['h2', 'h3', 'a'], class_=re.compile('name|title|company'))
                    if name:
                        exhibitor = {'name': name.get_text(strip=True)}
                        
                        # Try to find additional info
                        link = elem.find('a', href=True)
                        if link:
                            exhibitor['url'] = urljoin(url, link['href'])
                        
                        exhibitors.append(exhibitor)
                except:
                    continue
            
            print(f"   Found {len(exhibitors)} from category search")
            
        except Exception as e:
            print(f"   ⚠ Category scraping failed: {e}")
        
        return exhibitors
    
    def _scrape_by_directory(self):
        """Scrape alphabetical directory"""
        exhibitors = []
        
        # Try a few letters to get sample
        for letter in ['a', 'b', 'e', 'k', 'm', 's']:
            try:
                url = f"https://www.k-online.com/vis/v1/en/directory/{letter}"
                response = self.session.get(url, timeout=15)
                soup = BeautifulSoup(response.content, 'html.parser')
                
                # Find company listings
                companies = soup.find_all(['li', 'div'], class_=re.compile('company|exhibitor|entry'))
                
                for company in companies[:50]:  # Limit per letter
                    name_elem = company.find(['a', 'span', 'h3'])
                    if name_elem:
                        exhibitors.append({
                            'name': name_elem.get_text(strip=True),
                            'url': name_elem.get('href', '')
                        })
                
                time.sleep(0.5)  # Rate limiting
                
            except:
                continue
        
        print(f"   Found {len(exhibitors)} from directory")
        return exhibitors
    
    def _format_exhibitors(self, cached_data):
        """Format cached exhibitor data"""
        return [
            {
                'name': row[1],
                'url': row[2],
                'hall': row[3],
                'stand': row[4],
                'country': row[5],
                'products': json.loads(row[6]) if row[6] else []
            }
            for row in cached_data
        ]
    
    def get_fallback_exhibitors(self):
        """Return comprehensive list of major machinery manufacturers"""
        return [
            # Top Tier - Premium
            {'name': 'ENGEL Austria GmbH', 'country': 'Austria', 'tier': 'premium'},
            {'name': 'Arburg GmbH + Co KG', 'country': 'Germany', 'tier': 'premium'},
            {'name': 'KraussMaffei Technologies GmbH', 'country': 'Germany', 'tier': 'premium'},
            {'name': 'Sumitomo (SHI) Demag Plastics Machinery', 'country': 'Germany', 'tier': 'premium'},
            
            # Mid-Tier - Quality
            {'name': 'Husky Injection Molding Systems', 'country': 'Canada', 'tier': 'mid'},
            {'name': 'Wittmann Battenfeld GmbH', 'country': 'Austria', 'tier': 'mid'},
            {'name': 'Negri Bossi SpA', 'country': 'Italy', 'tier': 'mid'},
            {'name': 'Milacron LLC', 'country': 'USA', 'tier': 'mid'},
            {'name': 'BOY Machines Inc', 'country': 'Germany', 'tier': 'mid'},
            {'name': 'Nissei Plastic Industrial Co', 'country': 'Japan', 'tier': 'mid'},
            
            # Budget Tier
            {'name': 'Haitian International Holdings', 'country': 'China', 'tier': 'budget'},
            {'name': 'Chen Hsong Holdings', 'country': 'Hong Kong', 'tier': 'budget'},
            {'name': 'Borch Machinery', 'country': 'China', 'tier': 'budget'},
            {'name': 'Yizumi Precision Machinery', 'country': 'China', 'tier': 'budget'},
            
            # Specialized
            {'name': 'Sacmi Imola', 'country': 'Italy', 'tier': 'mid', 'specialty': 'Compression molding'},
            {'name': 'Netstal Maschinen AG', 'country': 'Switzerland', 'tier': 'premium', 'specialty': 'PET'},
            {'name': 'Sandretto Industrie', 'country': 'Italy', 'tier': 'mid'},
            {'name': 'Tederic Machinery', 'country': 'Taiwan', 'tier': 'budget'},
            {'name': 'Fu Chun Shin Machinery', 'country': 'Taiwan', 'tier': 'budget'},
            {'name': 'Windsor Machines', 'country': 'India', 'tier': 'budget'},
        ]


class FastMachineryMatcher:
    """Optimized matcher for large-scale analysis"""
    
    def __init__(self, api_key, cache_db):
        self.client = anthropic.Anthropic(api_key=api_key)
        self.cache = cache_db
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })
    
    def analyze_prospects_batch(self, prospects_df, enable_scraping=False):
        """Analyze prospects in batches for efficiency"""
        
        print("\n" + "="*90)
        print("📊 ANALYZING PROSPECTS")
        print("="*90)
        
        enriched = []
        total = len(prospects_df)
        
        # Process in batches
        batch_size = 10 if enable_scraping else 50
        
        for i in range(0, total, batch_size):
            batch = prospects_df.iloc[i:i+batch_size]
            print(f"\nBatch {i//batch_size + 1}/{(total + batch_size - 1)//batch_size}")
            
            for _, row in batch.iterrows():
                company = row.get('Firma', '')
                website = row.get('Web1', '')
                
                if not company:
                    continue
                
                # Check cache first
                cached = self.cache.get_prospect_cache(website) if website else None
                
                if cached:
                    enriched.append(cached)
                    print(f"  ✓ {company} (cached)")
                else:
                    # Analyze prospect
                    prospect_data = {
                        'name': company,
                        'country': row.get('Jud', ''),
                        'revenue_2024': float(row.get('Cifra2024EUR', 0)) if pd.notna(row.get('Cifra2024EUR')) else 0,
                        'website': website
                    }
                    
                    # Optional: detect machinery
                    if enable_scraping and website and website != '-':
                        machinery = self._quick_detect_machinery(company, website)
                        if machinery:
                            prospect_data['existing_machinery'] = machinery
                    
                    enriched.append(prospect_data)
                    
                    # Cache it
                    if website:
                        self.cache.save_prospect_cache(website, company, prospect_data)
                    
                    print(f"  ✓ {company}")
                    time.sleep(0.3)  # Rate limiting
        
        return enriched
    
    def _quick_detect_machinery(self, company, url):
        """Fast machinery detection (text only, no images for speed)"""
        try:
            response = self.session.get(url, timeout=10)
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Remove unnecessary elements
            for tag in soup(["script", "style", "nav", "footer"]):
                tag.decompose()
            
            text = soup.get_text(separator=' ', strip=True)
            text = ' '.join(text.split())[:3000]  # Limit text
            
            # Quick keyword search for brands
            brands_found = []
            keywords = ['ENGEL', 'Arburg', 'KraussMaffei', 'Sumitomo', 'Demag', 
                       'Husky', 'Wittmann', 'Battenfeld', 'Haitian', 'Negri Bossi']
            
            for brand in keywords:
                if brand.lower() in text.lower():
                    brands_found.append(brand)
            
            if brands_found:
                return [{'brand': b, 'confidence': 'medium'} for b in brands_found]
            
        except:
            pass
        
        return None
    
    def smart_match_analysis(self, prospects, providers, top_n=10, technology_filter=None):
        """Use AI to match prospects with providers - returns FULL prospect lists"""
        
        print("\n" + "="*90)
        print("🤖 AI MATCHING ANALYSIS")
        if technology_filter:
            print(f"🎯 TECHNOLOGY FOCUS: {technology_filter.upper()}")
        print("="*90)
        
        # First pass: Use AI to identify which provider TYPES match which prospect TYPES
        print("🎯 Phase 1: Identifying provider-prospect patterns...")
        
        # Categorize prospects by size/region
        prospect_categories = self._categorize_prospects(prospects)
        
        # Get AI analysis of provider capabilities
        provider_profiles = self._analyze_provider_profiles(providers[:50], technology_filter)
        
        if not provider_profiles:
            print("⚠ No providers found matching the technology filter!")
            return None
        
        # Second pass: Match ALL prospects to ALL providers
        print(f"\n🎯 Phase 2: Matching ALL {len(prospects)} prospects to {len(provider_profiles)} providers...")
        
        all_matches = []
        
        for provider in provider_profiles:
            print(f"\n  📊 Analyzing {provider['name']}...")
            
            matched_prospects = []
            
            for prospect in prospects:
                score, reasons = self._calculate_match(prospect, provider, technology_filter)
                
                if score >= 50:  # Threshold for good match
                    matched_prospects.append({
                        'name': prospect['name'],
                        'country': prospect.get('country', ''),
                        'revenue': prospect.get('revenue_2024', 0),
                        'website': prospect.get('website', ''),
                        'production_processes': prospect.get('production_processes', []),
                        'existing_machinery': prospect.get('existing_machinery', []),
                        'match_score': score,
                        'match_reasons': reasons
                    })
            
            coverage_pct = (len(matched_prospects) / len(prospects) * 100) if prospects else 0
            
            all_matches.append({
                'provider': provider,
                'matched_prospects': matched_prospects,
                'coverage_pct': coverage_pct,
                'total_matched': len(matched_prospects)
            })
            
            print(f"     ✓ Matched {len(matched_prospects)} prospects ({coverage_pct:.1f}%)")
        
        # Sort by coverage
        all_matches.sort(key=lambda x: x['coverage_pct'], reverse=True)
        
        # Format for output
        results = {
            'total_prospects': len(prospects),
            'total_providers_analyzed': len(provider_profiles),
            'technology_filter': technology_filter,
            'top_providers': []
        }
        
        for rank, match in enumerate(all_matches[:top_n], 1):
            provider_result = {
                'rank': rank,
                'name': match['provider']['name'],
                'country': match['provider'].get('country', ''),
                'technologies': match['provider'].get('technologies', []),
                'coverage_pct': round(match['coverage_pct'], 1),
                'total_prospects_matched': match['total_matched'],
                'reasons': match['provider'].get('key_strengths', []),
                'ideal_for': match['provider'].get('ideal_for', ''),
                'matched_prospects_full_list': match['matched_prospects']  # FULL LIST
            }
            results['top_providers'].append(provider_result)
        
        return results
    
    def _categorize_prospects(self, prospects):
        """Categorize prospects by size and region"""
        categories = {
            'eu_large': [],
            'eu_medium': [],
            'eu_small': [],
            'non_eu_large': [],
            'non_eu_medium': [],
            'non_eu_small': []
        }
        
        eu_countries = ['DE', 'FR', 'IT', 'ES', 'PL', 'RO', 'NL', 'BE', 'AT', 'CZ', 'HU', 'PT', 
                       'SE', 'GR', 'DK', 'FI', 'SK', 'IE', 'HR', 'BG', 'LT', 'SI', 'LV', 'EE']
        
        for p in prospects:
            revenue = p.get('revenue_2024', 0)
            country = p.get('country', '')
            
            # Determine size
            if revenue >= 30_000_000:
                size = 'large'
            elif revenue >= 5_000_000:
                size = 'medium'
            else:
                size = 'small'
            
            # Determine region
            region = 'eu' if country in eu_countries else 'non_eu'
            
            key = f"{region}_{size}"
            if key in categories:
                categories[key].append(p)
        
        return categories
    
    def _analyze_provider_profiles(self, providers, technology_filter=None):
        """Use AI to analyze each provider's capabilities, optionally filtered by technology"""
        
        print("  📋 Analyzing provider capabilities with AI...")
        if technology_filter:
            print(f"     🎯 Focusing on {technology_filter} specialists...")
        
        profiles = []
        
        # Batch providers for efficiency
        for i in range(0, len(providers), 10):
            batch = providers[i:i+10]
            
            tech_context = ""
            if technology_filter:
                tech_context = f"\nIMPORTANT: Focus on providers that specialize in {technology_filter}. Prioritize those with strong capabilities in this technology."
            
            prompt = f"""Analyze these machinery providers and determine their ideal customer profiles.

PROVIDERS:
{json.dumps(batch, indent=1)}
{tech_context}

For EACH provider, return their profile in this format:
{{
  "name": "Provider Name",
  "country": "Country",
  "tier": "budget/mid/premium",
  "technologies": ["injection molding", "extrusion", etc],
  "ideal_revenue_range": "€X-Y million",
  "ideal_regions": ["EU", "Eastern Europe", etc],
  "processes": ["injection molding", "extrusion", etc],
  "company_size_focus": ["small", "medium", "large"],
  "key_strengths": ["strength1", "strength2", "strength3"],
  "ideal_for": "Brief description"
}}

Return as JSON array."""

            try:
                message = self.client.messages.create(
                    model="claude-sonnet-4-5-20250929",
                    max_tokens=2048,
                    temperature=0.3,
                    messages=[{"role": "user", "content": prompt}]
                )
                
                response_text = message.content[0].text
                json_match = re.search(r'\[.*\]', response_text, re.DOTALL)
                
                if json_match:
                    batch_profiles = json.loads(json_match.group())
                    
                    # Filter by technology if specified
                    if technology_filter:
                        filtered_batch = []
                        tech_keywords = TECHNOLOGY_KEYWORDS.get(technology_filter, [])
                        
                        for profile in batch_profiles:
                            provider_techs = profile.get('technologies', []) + profile.get('processes', [])
                            provider_techs_str = ' '.join(provider_techs).lower()
                            
                            # Check if provider supports this technology
                            if any(keyword.lower() in provider_techs_str for keyword in tech_keywords):
                                profiles.append(profile)
                                filtered_batch.append(profile)
                        
                        if filtered_batch:
                            print(f"     ✓ Found {len(filtered_batch)} {technology_filter} specialists in this batch")
                    else:
                        profiles.extend(batch_profiles)
                
                time.sleep(1)
                
            except Exception as e:
                print(f"    ⚠ Error analyzing batch: {e}")
                # Use basic profiles
                for p in batch:
                    profiles.append({
                        'name': p['name'],
                        'country': p.get('country', ''),
                        'tier': p.get('tier', 'mid'),
                        'technologies': ['general'],
                        'ideal_regions': ['EU'],
                        'key_strengths': ['Quality machinery'],
                        'ideal_for': 'General manufacturing'
                    })
        
        return profiles
    
    def _calculate_match(self, prospect, provider):
        """Calculate match score between prospect and provider"""
        
        score = 0
        reasons = []
        
        # Revenue matching
        prospect_revenue = prospect.get('revenue_2024', 0)
        provider_tier = provider.get('tier', 'mid')
        
        if provider_tier == 'premium' and prospect_revenue >= 30_000_000:
            score += 30
            reasons.append("Revenue matches premium tier")
        elif provider_tier == 'mid' and 5_000_000 <= prospect_revenue < 30_000_000:
            score += 30
            reasons.append("Revenue matches mid-range tier")
        elif provider_tier == 'budget' and prospect_revenue < 10_000_000:
            score += 30
            reasons.append("Revenue matches budget tier")
        elif provider_tier == 'mid':
            score += 15  # Mid-range can serve most
        
        # Geographic matching
        prospect_country = prospect.get('country', '')
        provider_country = provider.get('country', '')
        ideal_regions = provider.get('ideal_regions', [])
        
        eu_countries = ['DE', 'FR', 'IT', 'ES', 'PL', 'RO', 'NL', 'BE', 'AT', 'CZ', 'HU']
        
        if prospect_country == provider_country:
            score += 20
            reasons.append(f"Same country ({prospect_country})")
        elif prospect_country in eu_countries and 'EU' in ideal_regions:
            score += 15
            reasons.append("EU provider for EU prospect")
        elif 'Global' in ideal_regions:
            score += 10
        
        # Check existing machinery
        existing = prospect.get('existing_machinery', [])
        if existing:
            existing_brands = [m.get('brand', '') for m in existing if isinstance(m, dict)]
            if provider['name'] in ' '.join(existing_brands):
                score += 10
                reasons.append("Already customer (expansion opportunity)")
            elif any(brand in ['Haitian', 'Chen Hsong'] for brand in existing_brands):
                score += 15
                reasons.append("Has budget brand (upgrade opportunity)")
        
        # Minimum viable score
        if score < 50:
            score = max(score, 40)  # Give everyone a baseline chance
        
        return score, reasons[:3]  # Return top 3 reasons


def export_to_excel(results, output_file="machinery_partners_full_lists.xlsx"):
    """Export results with FULL prospect lists to Excel"""
    
    print(f"\n📊 Exporting to Excel: {output_file}")
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        
        # Sheet 1: Summary
        summary_data = []
        for provider in results['top_providers']:
            summary_data.append({
                'Rank': provider['rank'],
                'Provider Name': provider['name'],
                'Country': provider['country'],
                'Coverage %': provider['coverage_pct'],
                'Total Prospects': provider['total_prospects_matched'],
                'Ideal For': provider['ideal_for']
            })
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Format summary sheet
        worksheet = writer.sheets['Summary']
        worksheet.column_dimensions['A'].width = 8
        worksheet.column_dimensions['B'].width = 35
        worksheet.column_dimensions['C'].width = 15
        worksheet.column_dimensions['D'].width = 12
        worksheet.column_dimensions['E'].width = 15
        worksheet.column_dimensions['F'].width = 50
        
        # Header formatting
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Sheet 2+: One sheet per provider with FULL prospect list
        for provider in results['top_providers']:
            sheet_name = f"#{provider['rank']} {provider['name'][:25]}"  # Excel limit: 31 chars
            
            # Create prospect list
            prospect_data = []
            for prospect in provider['matched_prospects_full_list']:
                
                existing_machinery = prospect.get('existing_machinery', [])
                machinery_str = ', '.join([
                    m.get('brand', '') if isinstance(m, dict) else str(m) 
                    for m in existing_machinery
                ]) if existing_machinery else 'None detected'
                
                reasons_str = '; '.join(prospect.get('match_reasons', []))
                
                prospect_data.append({
                    'Company Name': prospect['name'],
                    'Country': prospect['country'],
                    'Revenue (EUR)': prospect['revenue'],
                    'Website': prospect['website'],
                    'Existing Machinery': machinery_str,
                    'Match Score': prospect.get('match_score', 0),
                    'Why Good Match': reasons_str
                })
            
            prospects_df = pd.DataFrame(prospect_data)
            prospects_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Format prospect sheet
            worksheet = writer.sheets[sheet_name]
            worksheet.column_dimensions['A'].width = 40  # Company name
            worksheet.column_dimensions['B'].width = 12  # Country
            worksheet.column_dimensions['C'].width = 15  # Revenue
            worksheet.column_dimensions['D'].width = 40  # Website
            worksheet.column_dimensions['E'].width = 30  # Existing machinery
            worksheet.column_dimensions['F'].width = 12  # Score
            worksheet.column_dimensions['G'].width = 60  # Why match
            
            # Header formatting
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', wrap_text=True)
            
            # Add provider info at top
            worksheet.insert_rows(1, 3)
            worksheet['A1'] = f"Provider: {provider['name']}"
            worksheet['A1'].font = Font(size=14, bold=True)
            worksheet['A2'] = f"Coverage: {provider['coverage_pct']}% ({provider['total_prospects_matched']} prospects)"
            worksheet['A3'] = f"Why Partner: {', '.join(provider['reasons'][:3])}"
            worksheet.merge_cells('A1:G1')
            worksheet.merge_cells('A2:G2')
            worksheet.merge_cells('A3:G3')
    
    print(f"✓ Excel file created: {output_file}")
    print(f"  - Summary sheet with all providers")
    print(f"  - {len(results['top_providers'])} detailed sheets (one per provider)")
    print(f"  - Full prospect lists with contact info and match reasons")


def export_to_json(results, output_file="machinery_partners_full_data.json"):
    """Export complete results to JSON"""
    
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(results, f, indent=2, ensure_ascii=False)
    
    print(f"✓ JSON file created: {output_file}")


def main():
    """Main execution for 1500+ prospects"""
    
    print("\n" + "="*90)
    print("🎯 SCALABLE MACHINERY MATCHER v2.0")
    print("   Optimized for 1500+ prospects & 1900+ K2025 exhibitors")
    print("="*90)
    
    # Setup
    api_key = ANTHROPIC_API_KEY or input("\nEnter Anthropic API key: ").strip()
    csv_file = CSV_FILE_PATH or input("Enter CSV path: ").strip()
    
    enable_scraping = input("\nDetect existing machinery from websites? (SLOW for 1500 prospects) [y/N]: ").strip().lower() == 'y'
    
    top_n = int(input(f"How many top providers? (default 10): ").strip() or "10")
    
    # Initialize
    print("\n🔧 Initializing...")
    cache_db = CacheDB()
    client = anthropic.Anthropic(api_key=api_key)
    
    # Test API
    try:
        client.messages.create(
            model="claude-sonnet-4-5-20250929",
            max_tokens=10,
            messages=[{"role": "user", "content": "test"}]
        )
        print("✓ API connected")
    except Exception as e:
        print(f"❌ API Error: {e}")
        return
    
    # Load prospects
    print(f"\n📁 Loading prospects from {csv_file}...")
    df = pd.read_csv(csv_file, encoding='utf-8-sig')
    df = df[df['Firma'].notna()]
    print(f"✓ Loaded {len(df)} prospects")
    
    # Limit if needed
    if len(df) > MAX_PROSPECTS_TO_ANALYZE:
        print(f"⚡ Analyzing first {MAX_PROSPECTS_TO_ANALYZE} for performance")
        df = df.head(MAX_PROSPECTS_TO_ANALYZE)
    
    # Scrape K2025 exhibitors
    k2025_scraper = K2025Scraper(cache_db)
    providers = k2025_scraper.scrape_all_exhibitors()
    
    # Fallback to curated list if scraping fails
    if len(providers) < 20:
        print("⚠ Using fallback provider list")
        providers = k2025_scraper.get_fallback_exhibitors()
    
    print(f"✓ {len(providers)} machinery providers loaded")
    
    # Analyze prospects
    matcher = FastMachineryMatcher(api_key, cache_db)
    enriched_prospects = matcher.analyze_prospects_batch(df, enable_scraping, tech_filter)
    
    if not enriched_prospects:
        print("\n❌ No prospects match the technology filter!")
        print("Try running without filter or enable web scraping for better detection.")
        return
    
    # Match analysis
    results = matcher.smart_match_analysis(enriched_prospects, providers, top_n, tech_filter)
    
    if results:
        # Display summary
        print("\n" + "="*90)
        print(f"📊 TOP {top_n} MACHINERY PROVIDERS")
        if tech_filter:
            print(f"🎯 FILTERED BY: {tech_filter.upper()}")
        print(f"   Analyzed: {len(enriched_prospects)} prospects")
        print("="*90)
        
        for p in results['top_providers']:
            print(f"\n🏆 #{p['rank']}: {p['name']}")
            print(f"   Technologies: {', '.join(p.get('technologies', ['General']))}")
            print(f"   Coverage: {p['coverage_pct']}% ({p['total_prospects_matched']} prospects)")
            print(f"   Ideal for: {p['ideal_for']}")
            print(f"   Top reasons: {', '.join(p['reasons'][:2])}")
        
        # Export to files
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        print("\n" + "="*90)
        print("💾 EXPORTING RESULTS")
        print("="*90)
        
        tech_suffix = f"_{tech_filter}" if tech_filter else ""
        excel_file = f"machinery_partners{tech_suffix}_{timestamp}.xlsx"
        json_file = f"machinery_partners{tech_suffix}_{timestamp}.json"
        
        export_to_excel(results, excel_file)
        export_to_json(results, json_file)
        
        print("\n" + "="*90)
        print("✅ ANALYSIS COMPLETE!")
        print("="*90)
        print(f"\n📁 Your files:")
        print(f"   1. {excel_file}")
        print(f"      → Summary sheet + {len(results['top_providers'])} detailed sheets")
        print(f"      → FULL prospect lists with all contact info")
        if tech_filter:
            print(f"      → FILTERED: Only {tech_filter} prospects & providers")
        print(f"      → Ready to share with machinery providers!")
        print(f"\n   2. {json_file}")
        print(f"      → Complete data in JSON format")
        print(f"\n   3. machinery_cache.db")
        print(f"      → Cached data for faster future runs")
        
        print(f"\n💼 Next Steps:")
        print(f"   1. Open {excel_file}")
        print(f"   2. Review each provider's sheet")
        if tech_filter:
            print(f"   3. Contact {tech_filter} machinery specialists")
            print(f"   4. Show them your {len(enriched_prospects)} {tech_filter} prospects!")
        else:
            print(f"   3. Contact providers with their specific prospect lists")
            print(f"   4. Show them exactly which companies they can reach through you!")
        
    cache_db.conn.close()


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n⚠ Interrupted")
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()